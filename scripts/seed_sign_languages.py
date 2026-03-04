"""
Seed MongoDB with 50 sign language entries from CEP_sign_language.xlsx
Sheet: "50 words for data collection"

Column mapping (Excel → MongoDB):
    Col 2 (หมวดหมู่)          → category
    Col 3 (คำไทย)            → titleThai
    Col 4 (คำอังกฤษ)          → titleEng
    Col 5 (label)             → label  (auto-generated as lowercase titleEng if formula)
    Col 6 (วิธีการทำท่าทาง)     → signMethod

Usage:
    uv run python scripts/seed_sign_languages.py
"""

import asyncio
import re
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime
from dotenv import load_dotenv
import os

load_dotenv()

MONGODB_URL = os.getenv("MONGODB_URL", "mongodb://localhost:27017")
# Ensure authSource is set for Docker MongoDB with authentication
if "authSource" not in MONGODB_URL and "@" in MONGODB_URL:
    separator = "&" if "?" in MONGODB_URL else "?"
    MONGODB_URL += f"{separator}authSource=admin"
DATABASE_NAME = os.getenv("DATABASE_NAME", "smart_glove")
EXCEL_FILE = "CEP_sign_language.xlsx"
SHEET_NAME = "50 words for data collection"
COLLECTION = "sign_languages"


def make_label(title_eng: str) -> str:
    """Convert titleEng to label format: lowercase, spaces → underscores."""
    label = title_eng.strip().lower()
    label = re.sub(r"['\"]", "", label)  # remove apostrophes/quotes
    label = re.sub(r"\s+", "_", label)   # spaces to underscores
    return label


async def seed():
    # Connect to MongoDB
    client = AsyncIOMotorClient(MONGODB_URL)
    db = client[DATABASE_NAME]
    collection = db[COLLECTION]

    # Read Excel — sheet 4
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    entries = []
    for row_idx in range(2, ws.max_row + 1):
        title_thai = ws.cell(row=row_idx, column=3).value
        if not title_thai or not str(title_thai).strip():
            continue

        title_thai = str(title_thai).strip()
        category = str(ws.cell(row=row_idx, column=2).value or "").strip()
        title_eng = str(ws.cell(row=row_idx, column=4).value or "").strip()
        sign_method = str(ws.cell(row=row_idx, column=6).value or "").strip()

        # Label: use explicit value if not a formula, otherwise generate from titleEng
        raw_label = ws.cell(row=row_idx, column=5).value
        if raw_label and not str(raw_label).startswith("="):
            label = str(raw_label).strip()
        else:
            label = make_label(title_eng)

        now = datetime.utcnow()
        entries.append({
            "titleThai": title_thai,
            "titleEng": title_eng,
            "label": label,
            "category": category,
            "signMethod": sign_method,
            "imageUrl": "",
            "created_at": now,
            "updated_at": now,
        })

    print(f"Found {len(entries)} entries from '{SHEET_NAME}'")

    # Check existing data
    existing = await collection.count_documents({})
    if existing > 0:
        print(f"\n⚠ Collection '{COLLECTION}' already has {existing} documents.")
        confirm = input("Delete existing and re-seed? (y/N): ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            client.close()
            return
        await collection.delete_many({})
        print(f"Deleted {existing} existing documents.\n")

    # Insert
    result = await collection.insert_many(entries)
    print(f"Inserted {len(result.inserted_ids)} entries\n")

    # Summary table
    print(f"{'No.':<4} {'category':<12} {'titleThai':<12} {'titleEng':<18} {'label'}")
    print("=" * 70)
    for i, e in enumerate(entries, 1):
        print(f"{i:<4} {e['category']:<12} {e['titleThai']:<12} {e['titleEng']:<18} {e['label']}")

    client.close()
    print(f"\nDone!")


if __name__ == "__main__":
    asyncio.run(seed())
